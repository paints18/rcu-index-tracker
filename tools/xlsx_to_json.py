#!/usr/bin/env python3
"""One-time migration: RCU Index Tracker .xlsx -> data/pets.json

This exists to bootstrap data/pets.json from the Google Sheets export. After the
initial import, data/pets.json is the source of truth and is edited by hand --
do NOT re-run this against a newer spreadsheet export expecting a clean merge,
because it will not preserve any hand edits made to the JSON since.

Usage:
    python tools/xlsx_to_json.py "path/to/RCU Index Tracker (Make a Copy).xlsx"

Requires: openpyxl
"""

import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

# Sheets that hold no pet rows and are not part of the index.
SKIP_SHEETS = {"Read Me"}

# Column layout of every pet sheet. Row 1 is a title/counter row, row 2 is the
# header, pet rows start at row 3.
HEADER_ROW = 2
FIRST_DATA_ROW = 3
COL_EGG, COL_NAME, COL_RARITY, COL_CLICKS = 0, 1, 2, 3
FIRST_VARIANT_COL = 4
LAST_VARIANT_COL = 8  # exclusive


def slugify(value):
    """Lowercase kebab-case, ASCII alnum only. Stable and idempotent."""
    text = str(value or "").strip().lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def category_id(sheet_name):
    """'World 1' -> 'world1', 'No Currency' -> 'no-currency'.

    The trailing-number join matches the slug format agreed for pet IDs.
    """
    return re.sub(r"-(\d+)$", r"\1", slugify(sheet_name))


def cell_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_sheet(worksheet):
    """Return (variant_labels, [pet dicts]) for one sheet."""
    header = next(
        worksheet.iter_rows(
            min_row=HEADER_ROW, max_row=HEADER_ROW,
            max_col=LAST_VARIANT_COL, values_only=True,
        )
    )
    variant_labels = [cell_text(h) for h in header[FIRST_VARIANT_COL:LAST_VARIANT_COL]]
    variant_labels = [v for v in variant_labels if v]

    pets = []
    for row in worksheet.iter_rows(
        min_row=FIRST_DATA_ROW, max_row=worksheet.max_row,
        max_col=LAST_VARIANT_COL, values_only=True,
    ):
        name = cell_text(row[COL_NAME])
        if not name:
            # Blank name means an unused template row, not a pet.
            continue

        # A variant cell that is blank means that variant does not exist for
        # this pet. TRUE/FALSE both mean it exists -- the boolean is the
        # spreadsheet owner's own progress, which we deliberately discard.
        available = []
        for offset, label in enumerate(variant_labels):
            if cell_text(row[FIRST_VARIANT_COL + offset]) is not None:
                available.append(slugify(label))

        pets.append({
            "name": name,
            "egg": cell_text(row[COL_EGG]),
            "rarity": cell_text(row[COL_RARITY]),
            "clicks": cell_text(row[COL_CLICKS]),
            "variants": available,
        })

    return variant_labels, pets


def build_slug(pet):
    """Slug is the pet name alone, so moving a pet between eggs or categories
    keeps its identity. Names are unique across the whole index; the collision
    check in main() refuses to write if that ever stops being true.
    """
    return slugify(pet["name"])


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)

    src = Path(sys.argv[1])
    if not src.exists():
        sys.exit("No such file: %s" % src)

    out_path = Path(__file__).resolve().parent.parent / "data" / "pets.json"
    workbook = openpyxl.load_workbook(src, data_only=True)

    all_variants = []
    categories = []
    slugs = Counter()

    for worksheet in workbook.worksheets:
        if worksheet.title in SKIP_SHEETS:
            continue

        variant_labels, pets = parse_sheet(worksheet)
        for label in variant_labels:
            entry = {"id": slugify(label), "label": label}
            if entry not in all_variants:
                all_variants.append(entry)

        cat_id = category_id(worksheet.title)
        for pet in pets:
            pet["slug"] = build_slug(pet)
            slugs[pet["slug"]] += 1

        categories.append({
            "id": cat_id,
            "label": worksheet.title,
            "pets": [
                {
                    "slug": p["slug"],
                    "name": p["name"],
                    "egg": p["egg"],
                    "rarity": p["rarity"],
                    "clicks": p["clicks"],
                    "variants": p["variants"],
                }
                for p in pets
            ],
        })

    collisions = {s: n for s, n in slugs.items() if n > 1}
    if collisions:
        sys.exit("Slug collisions (same pet name twice) -- refusing to write:\n  " + "\n  ".join(
            "%s (x%d)" % (s, n) for s, n in sorted(collisions.items())
        ))

    doc = {
        "schemaVersion": 1,
        "generatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "variants": all_variants,
        "categories": categories,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(doc, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )

    total_pets = sum(len(c["pets"]) for c in categories)
    total_ticks = sum(len(p["variants"]) for c in categories for p in c["pets"])
    print("Wrote %s" % out_path)
    print("  categories: %d" % len(categories))
    print("  pets:       %d" % total_pets)
    print("  checkboxes: %d" % total_ticks)
    for c in categories:
        print("    %-20s %4d pets" % (c["id"], len(c["pets"])))


if __name__ == "__main__":
    main()
